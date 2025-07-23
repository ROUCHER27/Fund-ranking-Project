from openpyxl import load_workbook

def generate_excel_output(return_df, fund_scale_df, fund_performance, mgrcomp_rank, result_df, contrast_df, df3):
    df_dict = {
        '1.产品区间收益率': return_df,
        '2.产品规模权重': fund_scale_df,
        '3.加权计算过程': fund_performance,
        '4.加权业绩排名': mgrcomp_rank,
        '5.综合得分': result_df,
        '6.低分名单': contrast_df,
        '7.名单比对': df3
    }
    excel_output_process(df_dict, './data/fof基金模版.xlsx', './output/FOF基金排名.xlsx')

def excel_output_process(df_dict, input_path, output_path):
    template_excel = load_workbook(input_path)
    
    col_offsets = [0, 0, 0, 0, 0, 0, 0]
    for i, (sheet_name, df) in enumerate(df_dict.items()):
        df = df.reset_index(drop=True)
        sheet = template_excel[sheet_name]
        for r_idx, row in enumerate(df.values, start=2):
            for c_idx, value in enumerate(row, start=1):
                col_offset = col_offsets[(c_idx - 1) % len(col_offsets)]
                sheet.cell(row=r_idx + 1, column=c_idx + col_offset, value=value)
    template_excel.save(output_path)